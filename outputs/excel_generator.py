"""
Generates the Engagement Costing Excel annexure.
4 sheets: Effort & Fee Model, Timeline (Gantt), Resource Plan, Payment Schedule.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from config import BLENDED_RATE_USD, OUTPUTS_PATH

# Protiviti brand colors
PROTIVITI_RED = "C8102E"
DARK_GRAY = "404040"
LIGHT_GRAY = "F2F2F2"
MID_GRAY = "D9D9D9"
WHITE = "FFFFFF"
PHASE_COLORS = ["D6E4F0", "D5F5E3", "FEF9E7", "FDEDEC", "EAF2FF"]

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def generate_excel(effort_model: dict, client_name: str, project_title: str) -> str:
    """
    Generate the full costing Excel file.
    Returns the file path.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    _build_effort_sheet(wb, effort_model, client_name, project_title)
    _build_gantt_sheet(wb, effort_model)
    _build_resource_sheet(wb, effort_model)
    _build_payment_sheet(wb, effort_model, client_name)

    filename = f"Engagement_Costing_{client_name.replace(' ', '_')}.xlsx"
    output_path = os.path.join(OUTPUTS_PATH, filename)
    wb.save(output_path)
    return output_path


def _build_effort_sheet(wb, effort_model: dict, client_name: str, project_title: str):
    ws = wb.create_sheet("Effort & Fee Model")
    ws.freeze_panes = "A6"

    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = f"ENGAGEMENT COSTING MODEL — {project_title.upper()}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=PROTIVITI_RED)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Client: {client_name}  |  Blended Rate: USD {BLENDED_RATE_USD}/hour  |  Confidential"
    ws["A2"].font = Font(name="Calibri", size=10, color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=DARK_GRAY)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["Phase", "L1 Deliverable", "L2 Deliverable",
               "Hours", "Rate (USD)", "Fee (USD)", "% of Total",
               "Complexity", "Week Start", "Week End"]
    col_widths = [25, 30, 35, 10, 12, 14, 10, 12, 11, 11]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=PROTIVITI_RED)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 30

    total_fee = effort_model.get("total_fee_usd", 0)
    row = 5
    phases = effort_model.get("phases", [])

    for p_idx, phase in enumerate(phases):
        phase_color = PHASE_COLORS[p_idx % len(PHASE_COLORS)]
        phase_name = phase.get("phase_name", "")

        for deliv in phase.get("deliverables", []):
            l1_name = deliv.get("l1_name", "")
            first_l2 = True

            for l2 in deliv.get("sub_deliverables", []):
                ws.cell(row=row, column=1, value=phase_name if first_l2 else "")
                ws.cell(row=row, column=2, value=l1_name if first_l2 else "")
                ws.cell(row=row, column=3, value=l2.get("l2_name", ""))
                ws.cell(row=row, column=4, value=l2.get("hours", 0))
                ws.cell(row=row, column=5, value=BLENDED_RATE_USD)
                ws.cell(row=row, column=6, value=l2.get("fee_usd", 0))
                pct = (l2.get("fee_usd", 0) / total_fee * 100) if total_fee else 0
                ws.cell(row=row, column=7, value=round(pct, 1))
                ws.cell(row=row, column=8, value=l2.get("complexity", "medium").title())
                ws.cell(row=row, column=9, value=l2.get("week_start", ""))
                ws.cell(row=row, column=10, value=l2.get("week_end", ""))

                for col in range(1, 11):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill("solid", fgColor=phase_color)
                    cell.font = Font(name="Calibri", size=9)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if col in (4, 5, 6):
                        cell.number_format = '#,##0'
                    if col == 7:
                        cell.number_format = '0.0"%"'

                ws.row_dimensions[row].height = 18
                first_l2 = False
                row += 1

        # Phase subtotal row
        ws.cell(row=row, column=2, value=f"{phase_name} — SUBTOTAL")
        ws.cell(row=row, column=4, value=phase.get("phase_hours", 0))
        ws.cell(row=row, column=6, value=phase.get("phase_fee_usd", 0))
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=MID_GRAY)
            cell.font = Font(name="Calibri", size=9, bold=True)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if col in (4, 6):
                cell.number_format = '#,##0'
        ws.row_dimensions[row].height = 20
        row += 1

    # Grand total
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="TOTAL ENGAGEMENT")
    ws.cell(row=row, column=4, value=effort_model.get("total_hours", 0))
    ws.cell(row=row, column=5, value=BLENDED_RATE_USD)
    ws.cell(row=row, column=6, value=total_fee)
    ws.cell(row=row, column=7, value=100.0)
    for col in range(1, 11):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=PROTIVITI_RED)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        if col in (4, 6):
            cell.number_format = '#,##0'
    ws.row_dimensions[row].height = 25

    # Assumptions
    row += 3
    ws.cell(row=row, column=1, value="KEY ASSUMPTIONS")
    ws.cell(row=row, column=1).font = Font(bold=True, color=PROTIVITI_RED)
    for assumption in effort_model.get("assumptions", []):
        row += 1
        ws.cell(row=row, column=1, value=f"• {assumption}")
        ws.merge_cells(f"A{row}:J{row}")


def _build_gantt_sheet(wb, effort_model: dict):
    ws = wb.create_sheet("Execution Timeline")

    # Find max week
    max_week = 1
    for phase in effort_model.get("phases", []):
        for deliv in phase.get("deliverables", []):
            for l2 in deliv.get("sub_deliverables", []):
                max_week = max(max_week, l2.get("week_end", 1))

    max_week = max(max_week, 4)

    # Header
    ws.merge_cells(f"A1:{get_column_letter(3 + max_week)}1")
    ws["A1"] = "ENGAGEMENT TIMELINE — GANTT CHART"
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=PROTIVITI_RED)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    ws.cell(row=3, column=1, value="Phase").font = Font(bold=True)
    ws.cell(row=3, column=2, value="L1 Deliverable").font = Font(bold=True)
    ws.cell(row=3, column=3, value="L2 Deliverable").font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 32

    for w in range(1, max_week + 1):
        col = 3 + w
        cell = ws.cell(row=3, column=col, value=f"Wk {w}")
        cell.fill = PatternFill("solid", fgColor=DARK_GRAY)
        cell.font = Font(bold=True, color=WHITE, size=8)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 5

    row = 4
    for p_idx, phase in enumerate(effort_model.get("phases", [])):
        phase_color = PHASE_COLORS[p_idx % len(PHASE_COLORS)]
        bar_color = ["4472C4", "70AD47", "ED7D31", "FFC000", "5B9BD5"][p_idx % 5]

        for deliv in phase.get("deliverables", []):
            for l2 in deliv.get("sub_deliverables", []):
                ws.cell(row=row, column=1, value=phase.get("phase_name", ""))
                ws.cell(row=row, column=2, value=deliv.get("l1_name", ""))
                ws.cell(row=row, column=3, value=l2.get("l2_name", ""))

                w_start = l2.get("week_start", 1)
                w_end = l2.get("week_end", 1)

                for w in range(1, max_week + 1):
                    col = 3 + w
                    cell = ws.cell(row=row, column=col)
                    if w_start <= w <= w_end:
                        cell.fill = PatternFill("solid", fgColor=bar_color)
                        cell.value = "▓"
                        cell.font = Font(color=bar_color, size=7)
                    else:
                        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=phase_color)
                    ws.cell(row=row, column=col).font = Font(name="Calibri", size=8)
                    ws.cell(row=row, column=col).border = THIN_BORDER

                ws.row_dimensions[row].height = 16
                row += 1


def _build_resource_sheet(wb, effort_model: dict):
    ws = wb.create_sheet("Resource Plan")

    ws.merge_cells("A1:F1")
    ws["A1"] = "RESOURCE LOADING PLAN"
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=PROTIVITI_RED)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["Phase", "L1 Deliverable", "L2 Deliverable", "Hours", "Fee (USD)", "% of Total"]
    widths = [25, 30, 35, 12, 15, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_GRAY)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w

    total_fee = effort_model.get("total_fee_usd", 1)
    row = 4
    for p_idx, phase in enumerate(effort_model.get("phases", [])):
        phase_color = PHASE_COLORS[p_idx % len(PHASE_COLORS)]
        for deliv in phase.get("deliverables", []):
            for l2 in deliv.get("sub_deliverables", []):
                ws.cell(row=row, column=1, value=phase.get("phase_name", ""))
                ws.cell(row=row, column=2, value=deliv.get("l1_name", ""))
                ws.cell(row=row, column=3, value=l2.get("l2_name", ""))
                ws.cell(row=row, column=4, value=l2.get("hours", 0))
                ws.cell(row=row, column=5, value=l2.get("fee_usd", 0))
                pct = (l2.get("fee_usd", 0) / total_fee * 100)
                ws.cell(row=row, column=6, value=round(pct, 1))
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=phase_color)
                    ws.cell(row=row, column=col).border = THIN_BORDER
                    ws.cell(row=row, column=col).font = Font(name="Calibri", size=9)
                ws.row_dimensions[row].height = 16
                row += 1


def _build_payment_sheet(wb, effort_model: dict, client_name: str):
    ws = wb.create_sheet("Payment Schedule")

    ws.merge_cells("A1:F1")
    ws["A1"] = f"PAYMENT SCHEDULE — {client_name.upper()}"
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=PROTIVITI_RED)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Milestone", "Trigger / Deliverable", "Due Week", "Amount (USD)", "% of Total"]
    widths = [5, 30, 40, 12, 18, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_GRAY)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w

    total_fee = effort_model.get("total_fee_usd", 0)
    phases = effort_model.get("phases", [])

    # Auto-generate milestones: one per phase end
    row = 4
    cumulative = 0
    for i, phase in enumerate(phases, 1):
        phase_fee = phase.get("phase_fee_usd", 0)
        cumulative += phase_fee

        # Find last week for this phase
        last_week = 1
        for deliv in phase.get("deliverables", []):
            for l2 in deliv.get("sub_deliverables", []):
                last_week = max(last_week, l2.get("week_end", 1))

        pct = (phase_fee / total_fee * 100) if total_fee else 0
        trigger_deliverables = [d.get("l1_name", "") for d in phase.get("deliverables", [])]

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=f"Milestone {i}: {phase.get('phase_name', '')}")
        ws.cell(row=row, column=3, value=", ".join(trigger_deliverables[:3]))
        ws.cell(row=row, column=4, value=last_week)
        ws.cell(row=row, column=5, value=phase_fee)
        ws.cell(row=row, column=6, value=round(pct, 1))

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=PHASE_COLORS[i % len(PHASE_COLORS)])
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="center")
            if col == 5:
                cell.number_format = '$#,##0'
        ws.row_dimensions[row].height = 20
        row += 1

    # Total row
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=5, value=total_fee)
    ws.cell(row=row, column=6, value=100.0)
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=PROTIVITI_RED)
        cell.font = Font(bold=True, color=WHITE)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        if col == 5:
            cell.number_format = '$#,##0'
    ws.row_dimensions[row].height = 22
